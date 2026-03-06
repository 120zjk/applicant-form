import os
import json
import tempfile
import sys
from datetime import datetime
from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
import traceback
import random
import string
import filelock


def resource_path(relative_path):
    """Get the absolute path to a resource，works PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
# Configure Flask app
app = Flask(__name__,
            static_folder=resource_path('static'), #static files like css
            template_folder=resource_path('templates'))

app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB max file size
PHOTOS_ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'} # only for jpg, jpeg, png photo upload


# === Modified: Import custom logger configuration ===
from logger_config import setup_logging
logger = setup_logging(__name__)
# ====================================================


EXCEL_FILE = "HY-applicants.xlsx"# name of excel file
EXCEL_TEMPLATE_BACKUP = "HY-applicants-backup.xlsx"# backup file
EXCEL_LOCK = EXCEL_FILE + ".lock" #lock file
# store all submission records in memory for demo purpose
submission_records = []



def ensure_directories():
    """确保必要的目錄存在"""
    try:
        os.makedirs('templates', exist_ok=True)
        os.makedirs('static/css', exist_ok=True)
        logger.info("All directories are here")#log directory check complete,templates for store html files
    except Exception as e:
        logger.error(f"Directory creation error: {str(e)}")

def create_excel_template():
    """create a new Excel template file"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "basic info of applicant"
        #basic info sheet
        for col in["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z","AA", "AB", "AC", "AD", "AE", "AF"]:
            ws.column_dimensions[col].width = 45

        headers = ["Employee ID", "Type of form(frontline/general)","Position", "Salary", "Available Date", "Title",
                   "Chinese Name", "English Name", "ID/Passport Number", "Marital Status",
                   "Birth Date", "Birth Place", "Arrival Date", "Nationality","Race",
                   "Residential Phone", "Mobile Phone", "Email Address","Address","Correspondence Address", "Passport Issued Place",
                   "Passport Issued Date","Need Visa?","criminal offence","work in HY","relatives?","Submission Time"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')


        # education sheet
        ws_edu = wb.create_sheet(title="Education experience")

        for col in range(ord("A"), ord("Z") + 1):
            ws_edu.column_dimensions[chr(col)].width = 40

        education_headers = ["Employee ID", "type of form(general/frontline)", "Start Date", "Finish Date", "School/Institution","Qualification/Certificate obtained"]
        for col_idx, header in enumerate(education_headers, start=1):
            cell = ws_edu.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        #professional qualifications
        ws_prof = wb.create_sheet(title="Professional Qualification")
        for col in range(ord("A"), ord("Z") + 1):
            ws_prof.column_dimensions[chr(col)].width = 40
        
        prof_headers = ["Employee ID","type of form(general/frontline)", "Date Obtained", "Issuing Authority", "Qualification","License"]
        for col_idx, header in enumerate(prof_headers, start=1):
            cell = ws_prof.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # work sheet
        ws_work = wb.create_sheet(title="working experience")

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
            ws_work.column_dimensions[col].width = 20

        work_headers = ["Employee ID","formtype","Start Date", "Finish Date", "Name of Employer", "Previous Position", "Previous Salary","Reason of Leaving","Nature of Business"]
        for col_idx, header in enumerate(work_headers, start=1):
            cell = ws_work.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        #others sheet
        ws_others = wb.create_sheet(title="Others Information")
        for col in range(ord("A"), ord("Z") + 1):
            ws_others.column_dimensions[chr(col)].width = 40

        others_headers = ["Employee ID","formtype","Written in Chinese","Spoken in Chinese","Written in English","Spoken in English","Others language","Written in others","Spoken in others","Software Applications","Programming Languages","Others Skills",
                          "Company worked before","Department", "Position", "Duration",
                          "Name of relative working in company","Relationship with relative","Relative's company","Relative's department","Relative's position",
                          "criminal offence date","Criminal offence place",
                          "learn position",
                          "voluntary job-date","voluntary job union","voluntary job position"]

        for col_idx, header in enumerate(others_headers, start=1):
            cell = ws_others.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Referee sheet
        ws_ref = wb.create_sheet(title="Referee Information")
        for col in range(ord("A"), ord("Z") + 1):
            ws_ref.column_dimensions[chr(col)].width = 20

        referee_headers = ["Employee ID","formtype","Referee Name", "Phone Number", "Company", "Relationship"]

        for col_idx, header in enumerate(referee_headers, start=1):
            cell = ws_ref.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')


        wb.save(EXCEL_FILE)
        logger.info(f"Excel模板已創建: {EXCEL_FILE}")
        return wb
        
        os.system(f"cp {EXCEL_TEMPLATE} {EXCEL_TEMPLATE_BACKUP}") #create a backup

    except Exception as e:
        logger.error(f"FK error: {str(e)}")
        raise e

def load_excel_template():
    try:
        if not os.path.exists(EXCEL_FILE):
            logger.info("模版不在，正在創建新模版...")
            return create_excel_template()
        else:
            return openpyxl.load_workbook(EXCEL_FILE)

    except Exception as e:
        logger.error(f"Excel loading error: {str(e)}")
        logger.error(traceback.format_exc())
    

def get_next_employee_id():
    now = datetime.now() 
    timestamp_part = now.strftime("%Y%m%d%H%M%S")
    random_part = ''.join(random.choices(string.digits, k=3))# add three ramdom number at the end
    employee_id = f"{timestamp_part}{random_part}"
    return employee_id

def get_formType(form_data):
    try:
        formtype = form_data.get('formtype', '')
        if formtype in ['frontline', 'general']:
            return formtype
        else:
            logger.warning(f"未知的表單類型: {formtype}，默認設置為 'general'")
            return 'general'
    except Exception as e:
        logger.warning(f"error when getting form type: {str(e)}")
        return 'general'
    
def safe_set_cell_value(ws, cell_address, value):
    try:
        if value is not None and str(value).strip():
        
            cell = ws[cell_address]
            if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                logger.warning(f"跳过合并单元格: {cell_address}")
                return False
            
            ws[cell_address] = str(value).strip()
            return True
    except Exception as e:
        logger.warning(f"設置cell {cell_address} error: {str(e)}")
        return False
    
def analys_address_data(address1,address2,address3,address_area,address_district,address_country): # combine address parts into one line
    address = address1 + " " + address2 + " " + address3 + ", " + address_area + ", " + address_district + ", " + address_country
    return address

def fill_excel_with_form_data(form_data):
    """fill data into Excel"""
    try:
        logger.info(f"Starting to fill Excel data... applicants: {form_data.get('chinese_name', 'unknown')}")

        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
        else:
            wb = create_excel_template()
        
        # generate new Employee ID
        employee_id = get_next_employee_id()
        english_name = form_data.get('english_name', '')
        
        # ==================== main sheet ====================
        try:
            ws_main = wb["申請者基本資料"]
        except KeyError:
            try:
                ws_main = wb["basic info of applicant"]#新名
            except KeyError:
                ws_main = wb.active
                ws_main.title = "basic info of applicant"
        
        # 找到下一个空行
        next_row = ws_main.max_row + 1
        typeform = form_data.get('typeform', '')
        submission_time = datetime.now()
        
        
        # 基本資料
        main_data = [
            employee_id,  # Employee ID
            form_data.get('formtype', ''),  # 表單類型
            form_data.get('position', ''),  # 申請職位
            form_data.get('salary', ''),  # 要求薪金
            form_data.get('available_date', ''),  # 可到職日期
            form_data.get('title', ''),  # 稱謂
            form_data.get('chinese_name', ''),  # 中文姓名
            form_data.get('english_name', ''),  # 英文姓名
            form_data.get('id_number', ''),  # 身份證/護照號碼
            form_data.get('marital_status', ''),  # 婚姻狀況
            form_data.get('birth_date', ''),  # 出生日期
            form_data.get('birth_place', ''),  # 出生地點
            form_data.get('arrival_date', ''),  # 來港日期
            form_data.get('nationality', ''),  # 國籍
            form_data.get('race', ''),  # 種族
            form_data.get('home_phone', ''),  # 住宅電話
            form_data.get('mobile_phone', ''),  # 手提電話
            form_data.get('email', ''),  # 工作電話
            analys_address_data(form_data.get('address1',''),#地址
                                form_data.get('address2',''),
                                form_data.get('address3',''),
                                form_data.get('address_area',''),
                                form_data.get('address_district',''),
                                form_data.get('address_country','')
            ),
            form_data.get('correspondence_address', ''),  # 通訊地址
            form_data.get('passport_issue_place', ''),  # 護照簽發地點
            form_data.get('passport_issue_date', ''),  # 護照簽發日期
            form_data.get('visa_required', ''),  # 簽證/限制
            form_data.get('criminal_conviction', ''),  # 是否有犯罪記錄
            form_data.get('previous_employment', ''),  # 是否曾在康業工作
            form_data.get('has_relatives', ''),  # 是否有親屬在康業工作
            submission_time.strftime("%Y-%m-%d %H:%M:%S") # 提交時間
        ]
        
        # 寫入數據
        for col_idx, value in enumerate(main_data, start=1):
            safe_set_cell_value(ws_main, next_row, col_idx, value)
        
        # ==================== 填充教育經歷表 ====================
        
        try:
            ws_edu = wb["Education experience"]
        except KeyError:
            ws_edu = wb.create_sheet(title="Education experience")
            edu_headers = ["Employee ID", "Form Type", "Start Date", "Finish Date", "School/Institution", "Qualification/Certificate obtained"]
            for col_idx, header in enumerate(edu_headers, start=1):
                cell = ws_edu.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自動後一行資料填充
        edu_next_row = ws_edu.max_row + 1
        i = 1
        entries_added = 0
        while True:
            school = form_data.get(f'school_{i}', '')
            qualification = form_data.get(f'qualification_{i}', '')

            if not school and not qualification:
                break  # 如果學校和學歷都為空，停止loop

            edu_data = [
                    employee_id,
                    form_data.get('formtype', ''),
                    form_data.get(f'edu-start-date_{i}', ''),
                    form_data.get(f'edu-end-date_{i}', ''),
                    school,
                    qualification
                ]

            for col_idx, value in enumerate(edu_data, start=1):
                safe_set_cell_value(ws_edu, edu_next_row, col_idx, value)
                
            edu_next_row += 1
            i += 1
            entries_added += 1

            if entries_added > 0:
                logger.info(f"為 {english_name} 成功填充了 {entries_added} 筆教育資料")
        # =========================others sheet ==================================
        try:
            ws_others = wb["Others Information"]
        except KeyError:
            ws_others = wb.create_sheet(title="Others Information")
            for col in range(ord("A"), ord("Z") + 1):
                ws_others.column_dimensions[chr(col)].width = 40

                others_headers = ["Employee ID","formtype","Written in Chinese","Spoken in Chinese","Written in English","Spoken in English","Written in others","Spoken in others","Software Applications","Programming Languages","Others Skills",
                          "Company worked before","Department", "Position", "Duration",
                          "Name of relative working in company","Relationship with relative","Relative's company","Relative's department","Relative's position",
                          "criminal offence date","Criminal offence place",
                          "learn position",
                          "voluntary job-date","voluntary job union","voluntary job position"]

            for col_idx, header in enumerate(others_headers, start=1):
                cell = ws_others.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

        others_next_row = ws_others.max_row + 1
        
        other_data = [
            employee_id,
            form_data.get('formtype', ''),# type of form
            form_data.get('chinese_written', ''),#中文寫作
            form_data.get('chinese_spoken', ''),#中文口語
            form_data.get('english_written', ''),#英文寫作
            form_data.get('english_spoken', ''),#英文口語
            form_data.get('other_language',''),#其他語言
            form_data.get('other_written',''),#其他語言的寫作
            form_data.get('other_spoken',''),#其他語言的口語
            form_data.get('software_skills', ''),#軟件應用
            form_data.get('programming_skills', ''),#編程語言
            form_data.get('other_skills', ''),#其他技能
            #+++++++++++++++++++++++++++++++++++++++++++我只是一條分界線不要理會我+++++++++++++++++++++++++++++++++++++++++++++
            form_data.get('pre_employment1', ''),#以前工作公司
            form_data.get('pre_department1', ''),#以前工作部門
            form_data.get('pre_position1', ''),#以前工作職位
            form_data.get('pre_period1', ''),#the duration of previous job
            form_data.get('relative_name1', ''),#親屬姓名
            form_data.get('relative_relationship1', ''),#親屬關係
            form_data.get('relative_company1', ''),#親屬公司
            form_data.get('relative_department1', ''),#親屬部門
            form_data.get('relative_position1', ''),#親屬職位
            form_data.get('conviction_date', ''),#犯罪日期
            form_data.get('conviction_place', ''),#犯罪地點
            form_data.get('job_source', ''),#where u know the job
            form_data.get('voluntary-date', ''),#義工blablabal時間
            form_data.get('voluntary_organization', ''),#義工機構
            form_data.get('voluntary_position', '')#義工職位
        ]
        if any(v for v in other_data[2:]):
            for col_idx, value in enumerate(other_data, start=1):
                safe_set_cell_value(ws_others, others_next_row, col_idx, value)
            logger.info(f"Others information data filled for {english_name}")


        #=======================professional qualifications=======================
        try:
            ws_prof = wb["Professional Qualification"]
        except KeyError:
            ws_prof = wb.create_sheet(title="Professional Qualification")
            for col in range(ord("A"), ord("Z") + 1):
                ws_prof.column_dimensions[chr(col)].width = 20
                prof_headers = ["Employee ID","type of form", "Date Obtained", "Issuing Authority", "Qualification","License"]
            for col_idx, header in enumerate(prof_headers, start=1):
                cell = ws_prof.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        prof_next_row = ws_prof.max_row + 1
        i = 1
        while True:
            prof_data = [
                employee_id,  # Employee ID
                form_data.get('formtype', ''),  # type of form
                form_data.get(f'prof_date_obtained_{i}', ''),  # Date Obtained
                form_data.get(f'institution_{i}', ''),  # Issuing Authority
                form_data.get(f'pro_qualification_{i}', ''),  # Qualification
                form_data.get(f'prof_license_{i}', '')  # License
            ]
                
            for col_idx, value in enumerate(prof_data, start=1):
                safe_set_cell_value(ws_prof, prof_next_row, col_idx, value)
            prof_next_row += 1
            i += 1
            if not any(prof_data[2:]):
                break
        logger.info(f"Professional qualification data filled for {english_name}, entries: {i-1}")
    #============================== working experience =========================
        try:
            ws_work = wb["working experience"]
        except KeyError:
            ws_work = wb.create_sheet(title="working experience")
            for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
                ws_work.column_dimensions[col].width = 40
            work_headers = ["Employee ID","formtype","Start Date", "Finish Date", "Name of Employer", "Previous Position", "Previous Salary","Reason of Leaving","Nature of Business"]
            for col_idx, header in enumerate(work_headers, start=1):
                cell = ws_work.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        work_next_row = ws_work.max_row + 1
        i = 1
        while True:
            work_data = [
                employee_id,  # Employee ID
                form_data.get('formtype', ''),  # type of form
                form_data.get(f'work-start{i}', ''),  # Start Date
                form_data.get(f'work-end{i}', ''),  # Finish Date
                form_data.get(f'employer_{i}', ''),  # Name of Employer
                form_data.get(f'last_position_{i}', ''),  # Previous Position
                form_data.get(f'last_salary_{i}', ''),  # Previous Salary
                form_data.get(f'reason_leaving_{i}', ''),  # Reason of Leaving
                form_data.get(f'business_nature_{i}', '')  # Nature of Business
                ]
            for col_idx, value in enumerate(work_data, start=1):
                safe_set_cell_value(ws_work, work_next_row, col_idx, value) 
            work_next_row += 1
            i += 1
            if not any(work_data[3:]):
                break
        logger.info(f"Working experience data filled for {english_name}, entries: {i-1}")
        
        # ==================== 填充諮詢人資料表 ====================
        try:
            ws_ref = wb["Referee Information"]
        except KeyError:
            # 创建Referee Information工作表
            ws_ref = wb.create_sheet(title="Referee Information")
            ref_headers = ["Employee ID", "type of form", "Referee Name", "Phone Number", "Address", "Relationship"]
            for col_idx, header in enumerate(ref_headers, start=1):
                cell = ws_ref.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="A349A4", end_color="A349A4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ref_next_row = ws_ref.max_row + 1
        # Referee 1
        if form_data.get('referee1_name'):
            ref1_data = [
                employee_id,
                form_data.get('formtype', ''),
                form_data.get('referee1_name'),
                form_data.get('referee1_position'),
                form_data.get('referee1_contact'),
                form_data.get('referee1_company')
            ]
            for col_idx, value in enumerate(ref1_data, start=1):
                safe_set_cell_value(ws_ref, ref_next_row, col_idx, value)
            ref_next_row += 1

        # Referee 2
        if form_data.get('referee2_name'):
            ref2_data = [
                employee_id,
                form_data.get('formtype', ''),
                form_data.get('referee2_name'),
                form_data.get('referee2_position'),
                form_data.get('referee2_contact'),
                form_data.get('referee2_company')
            ]
            for col_idx, value in enumerate(ref2_data, start=1):
                safe_set_cell_value(ws_ref, ref_next_row, col_idx, value)
            ref_next_row += 1
            logger.info(f"Referee data filled for {english_name}")

        
        
        # 保存更新
        wb.save(EXCEL_FILE)
        
        # record submission
        submission_record = {
            'id': len(submission_records) + 1,
            'employee_id': employee_id,
            'chinese_name': form_data.get('chinese_name', ''),
            'english_name': english_name,
            'position': form_data.get('position', ''),
            'submission_time': submission_time.isoformat(),
            'status': 'completed'
        }
        submission_records.append(submission_record)
        
        return submission_record
        
    except Exception as e:
        logger.error(f"fill excel data error: {str(e)}")
        logger.error(traceback.format_exc())
        raise

def safe_set_cell_value(ws, row, col, value):
    # 安全設置單元格值
    try:
        if value is not None and str(value).strip():
            cell = ws.cell(row=row, column=col)
            cell.value = str(value).strip()
            return True
        return False
    except Exception as e:
        logger.warning(f"設置單元格 行{row}列{col} 時出錯: {str(e)}")
        return False


@app.route('/')
def index():
    """Home route"""
    try:
        logger.info(f"visit - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        return render_template('index.html')
    except Exception as e:
        logger.error(f"Error-from index: {str(e)}")
        return f"Error-from index: {str(e)}", 500

@app.route('/form1')
def form1():
    """form1 (the frontline position) route"""
    try:
        logger.info(f"user HY_hr visit form1 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        return render_template('form1.html')
    except Exception as e:
        logger.error(f"Error-from form1: {str(e)}")
        return f"Error-from form1: {str(e)}", 500
    
@app.route('/form2')
def form2():
    """form2 (the non-frontline position) route"""
    try:
        logger.info(f"user HY_hr visit form2 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        return render_template('form2.html')
    except Exception as e:
        logger.error(f"Error-from form2: {str(e)}")
        return f"Error-from form2: {str(e)}", 500
    
@app.route('/success')
def success():
    """提交成功"""
    try:
        logger.info(f"user HY_hr visit success - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        return render_template('success.html')
    except Exception as e:
        logger.error(f"Error-from success: {str(e)}")
        return f"Error-from success: {str(e)}", 500

@app.route('/submit', methods=['POST'])
def submit_form():
    """提交表單並填充Excel"""
    try:
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        logger.info(f"visit - {current_time}")

        # 檢查頁面表單
        if not request.form:
            logger.error("empty data from web")
            return jsonify({
                'success': False,
                'error': 'no form data'
            }), 400
        
        # 分析logger
        form_data = {}
        for key in request.form.keys():
            try:
                value = request.form.get(key, '').strip()
                form_data[key] = value
                if value:  
                    logger.debug(f"表单字段 {key}: {value[:50]}...")  # 限制日志长度
            except Exception as e:
                logger.warning(f"处理表单字段 {key} 时出错: {str(e)}")
                form_data[key] = ''
        
        # 确保基本字段存在
        chinese_name = form_data.get('chinese_name', '').strip()
        position = form_data.get('position', '').strip()
        
        if not chinese_name:
            logger.error("缺少中文姓名")
            return jsonify({
                'success': False,
                'error': '請填寫中文姓名'
            }), 400
            
        if not position:
            logger.error("缺少申请职位")
            return jsonify({
                'success': False,
                'error': '請填寫申請職位'
            }), 400
        
        logger.info(f"處理申請人: {chinese_name} 申請職位: {position}")
        
        # 填充Excel
        submission_record = fill_excel_with_form_data(form_data)
        
    
    except Exception as e:
        logger.error(f"處理表單時出錯: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f"處理表單時出錯，請重試。錯誤詳情: {str(e)}"
        }), 500



@app.route('/health')
def health_check():
    return "OK"

@app.errorhandler(404)
def not_found(error):
    logger.warning(f"404 error: {request.url}")
    return jsonify({
        'success': False,
        'error': '找不到請求的資源'
    }), 404

@app.errorhandler(500) 
def internal_error(error):
    logger.error(f"500 error: {str(error)}")
    return jsonify({
        'success': False,
        'error': '服務器內部錯誤'
    }), 500

if __name__ == '__main__':
    try:
        # 確保目錄存在
        ensure_directories()
        logger.info(f"康業控股求職申請系統啟動 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        app.run(debug=False, host='0.0.0.0', port=5000)
        
    except Exception as e:
        logger.error(f"Error-from main: {str(e)}")
        logger.error(traceback.format_exc())